#!/usr/bin/env python3
"""
sync_contentores.py
Sincroniza o Excel CONTENTORES-MATERIAL Final-2026.xlsx para o Google Sheet,
preservando TODA a formatação, cores, hiperligações e abas.

Estratégia:
  1. Converte o Excel num Google Sheet temporário (via Drive API)
  2. Copia todas as abas do temporário para o ficheiro permanente
  3. Apaga as abas antigas do ficheiro permanente
  4. Apaga o ficheiro temporário

Dependências:
  pip install openpyxl gspread google-api-python-client google-auth
"""

import os
import time
import openpyxl
from datetime import datetime, date
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

MESES_PT = ["janeiro","fevereiro","março","abril","maio","junho",
            "julho","agosto","setembro","outubro","novembro","dezembro"]

DATE_COLS_FC = [6, 7, 9]   # G (ETD), H (ETA), J (entrega) — I é PRONTO (SIM/NÃO)
TAB_FC       = "Fornecedores_Contentores"
EXCEL_TAB_FC = "Fornecedores_Contentores"

def fmt_date_pt(v):
    if isinstance(v, (datetime, date)):
        return f"{v.day} de {MESES_PT[v.month - 1]} de {v.year}"
    return None

# ── Config ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "CONTENTORES-MATERIAL Final-2026.xlsx")
CREDS_FILE = os.path.join(SCRIPT_DIR, "google_credentials.json")
SHEET_ID   = "1cfqw7GGGzk5GQOUcVluIrQOUySlPDoAwch4MrwkSkBA"

SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
]

TEMP_NAME = "__sync_temp__"
OLD_PREFIX = "__old__"

# ── Helpers ───────────────────────────────────────────────────────────────────

def log(msg):
    print("[" + datetime.now().strftime("%H:%M:%S") + "] " + msg)

def get_services():
    creds = service_account.Credentials.from_service_account_file(
        CREDS_FILE, scopes=SCOPES
    )
    drive  = build("drive",  "v3", credentials=creds)
    sheets = build("sheets", "v4", credentials=creds)
    return drive, sheets

# ── Limpeza de runs anteriores falhados ──────────────────────────────────────

def cleanup_old_prefix(sheets_svc):
    """Remove abas com prefixo __old__ que ficaram de runs anteriores."""
    info = sheets_svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    reqs = []
    for s in info["sheets"]:
        if s["properties"]["title"].startswith(OLD_PREFIX):
            reqs.append({"deleteSheet": {"sheetId": s["properties"]["sheetId"]}})
    if reqs:
        log("A limpar abas residuais de runs anteriores...")
        sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=SHEET_ID, body={"requests": reqs}
        ).execute()

def cleanup_temp_files(drive_svc):
    """Remove TODOS os ficheiros pertencentes à conta de serviço (liberta quota)."""
    page_token = None
    total = 0
    while True:
        kwargs = {
            "q": "'me' in owners and trashed=false",
            "fields": "nextPageToken,files(id,name)",
            "pageSize": 100,
        }
        if page_token:
            kwargs["pageToken"] = page_token
        result = drive_svc.files().list(**kwargs).execute()
        files = result.get("files", [])
        for f in files:
            try:
                drive_svc.files().delete(fileId=f["id"]).execute()
                total += 1
                log("Apagado: " + f["name"])
            except Exception as e:
                log("Aviso ao apagar " + f["name"] + ": " + str(e))
        page_token = result.get("nextPageToken")
        if not page_token:
            break
    if total:
        log("Limpeza concluída: " + str(total) + " ficheiros eliminados.")

# ── Sync principal ────────────────────────────────────────────────────────────

def validate_excel(path):
    """Verifica que o ficheiro Excel é um ZIP válido com dados suficientes."""
    import zipfile
    try:
        with zipfile.ZipFile(path, 'r') as z:
            names = z.namelist()
            if not any(n.startswith("xl/") for n in names):
                return False, "Estrutura xlsx inválida (sem pasta xl/)"
    except zipfile.BadZipFile as e:
        return False, "Ficheiro corrompido: " + str(e)
    # Verificar número mínimo de contentores
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if EXCEL_TAB_FC not in wb.sheetnames:
            return False, "Tab " + EXCEL_TAB_FC + " não encontrada"
        ws = wb[EXCEL_TAB_FC]
        count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[3])
        wb.close()
        if count < 10:
            return False, f"Apenas {count} contentores — possível versão incompleta"
    except Exception as e:
        return False, "Erro ao ler Excel: " + str(e)
    return True, f"OK ({count} contentores)"

def sync(drive_svc, sheets_svc):
    """
    Substitui o conteúdo do Google Sheet directamente com o Excel.
    Não cria ficheiros novos — evita o problema de quota da conta de serviço.
    O Drive converte automaticamente o xlsx para formato Sheets.
    """
    # Validar ficheiro antes de fazer upload
    ok, msg = validate_excel(EXCEL_FILE)
    if not ok:
        log("ERRO: Sync abortado — " + msg)
        log("O Google Sheet NÃO foi alterado.")
        return
    log("Ficheiro validado: " + msg)
    log("A substituir conteúdo do Google Sheet com o Excel...")

    media = MediaFileUpload(
        EXCEL_FILE,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True
    )

    result = drive_svc.files().update(
        fileId=SHEET_ID,
        media_body=media,
        supportsAllDrives=True,
        fields="id,name,mimeType"
    ).execute()

    log("Ficheiro actualizado: " + result.get("name", SHEET_ID))
    log("MimeType: " + result.get("mimeType", "?"))

    # Verificar se o Drive manteve o formato Sheets
    if result.get("mimeType") != "application/vnd.google-apps.spreadsheet":
        log("AVISO: O ficheiro foi convertido para xlsx em vez de Sheets.")
        log("Consulta o README ou pede ajuda para resolver.")
    else:
        # Aguardar que o Drive propague as alterações antes de usar Sheets API
        log("A aguardar propagação (5s)...")
        time.sleep(5)
        # Contar abas sincronizadas
        info = sheets_svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
        n = len(info.get("sheets", []))
        log("Sincronização concluída! " + str(n) + " abas actualizadas.")
        # Corrigir datas em português na tab Fornecedores_Contentores
        fix_dates_pt(sheets_svc)


def fix_dates_pt(sheets_svc):
    """Reescreve colunas G/H/I/J de Fornecedores_Contentores com datas em português."""
    log("A corrigir formato de datas (português)...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb[EXCEL_TAB_FC]
    except Exception as e:
        log("AVISO: Não foi possível ler Excel para datas: " + str(e))
        return

    # Obter sheetId da tab
    info = sheets_svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
    sheet_id = None
    for s in info["sheets"]:
        if s["properties"]["title"] == TAB_FC:
            sheet_id = s["properties"]["sheetId"]
            break
    if sheet_id is None:
        log("AVISO: Tab " + TAB_FC + " não encontrada no Sheets.")
        return

    requests = []
    for row_i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        for col_i in DATE_COLS_FC:
            if col_i >= len(row):
                continue
            cell = row[col_i]
            pt_str = fmt_date_pt(cell.value)
            if pt_str:
                requests.append({
                    "updateCells": {
                        "range": {
                            "sheetId":          sheet_id,
                            "startRowIndex":    row_i,
                            "endRowIndex":      row_i + 1,
                            "startColumnIndex": col_i,
                            "endColumnIndex":   col_i + 1,
                        },
                        "rows": [{"values": [{"userEnteredValue": {"stringValue": pt_str}}]}],
                        "fields": "userEnteredValue",
                    }
                })

    if requests:
        sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=SHEET_ID, body={"requests": requests}
        ).execute()
        log("Datas corrigidas: " + str(len(requests)) + " células.")
    else:
        log("Nenhuma data a corrigir.")

# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    log("=" * 52)
    log("  Sync Excel → Google Sheets  " + datetime.now().strftime("%Y-%m-%d %H:%M"))
    log("=" * 52)

    if not os.path.exists(EXCEL_FILE):
        log("ERRO: Ficheiro Excel não encontrado: " + EXCEL_FILE)
        return
    if not os.path.exists(CREDS_FILE):
        log("ERRO: Credenciais não encontradas: " + CREDS_FILE)
        return

    drive_svc, sheets_svc = get_services()

    # Limpeza de possíveis resíduos de runs anteriores
    cleanup_temp_files(drive_svc)
    cleanup_old_prefix(sheets_svc)

    sync(drive_svc, sheets_svc)

if __name__ == "__main__":
    main()
